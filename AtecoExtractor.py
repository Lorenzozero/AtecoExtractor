import tkinter as tk
from tkinter import messagebox, ttk
import requests
from urllib.parse import urljoin
import parsel
import openpyxl
import mysql.connector
import csv
import threading

def extract_fatturato_from_page(page_content):
    selector = parsel.Selector(text=page_content)
    fatturato_links = selector.css('.col-8.col-lg-12.content-azienda a[title^="Bilancio"]::text').getall()

    if not hasattr(extract_fatturato_from_page, "used_indices"):
        extract_fatturato_from_page.used_indices = set()

    for i, link in enumerate(fatturato_links):
        if i not in extract_fatturato_from_page.used_indices:
            extract_fatturato_from_page.used_indices.add(i)
            # Rimuovi il simbolo euro dalla stringa del fatturato
            fatturato_value = link.strip().replace('€', '')
            return fatturato_value

    extract_fatturato_from_page.used_indices = set()
    return "Non disponibile"

def scrape_company_data(url, ateco_value, progress_var, active_text, inactive_text, active_count_var, inactive_count_var):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Solleva un'eccezione per errori HTTP
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Errore", f"Errore durante la richiesta HTTP: {str(e)}")
        return None

    selector = parsel.Selector(text=response.text)

    # Estrazione delle parole in grassetto all'interno del paragrafo
    bold_texts = selector.css('p.text-justify > b::text').getall()

    # Estrazione del testo completo all'interno del paragrafo text-justify
    paragraph_text = selector.css('div.col-12::text').get()

    # Estrazione dei valori delle parole in grassetto
    if len(bold_texts) >= 3:
        company_name = bold_texts[0].strip()
        company_address = bold_texts[1].strip()
        company_code_rea = bold_texts[2].strip()
    else:
        messagebox.showwarning("Informazioni non trovate", "Informazioni non trovate")
        return None

    # Estrazione della Partita IVA
    partita_iva = selector.xpath("//p[contains(., 'Partita IVA')]/b/text()").get()

    # Estrazione dello Stato Attività
    stato_attivita = selector.xpath("//div[@class='col-5'][p/b='Stato Attività']/following-sibling::div[@class='col-7']/p/text()").get()

    # Controllo se lo "Stato Attività" è "Attiva" prima di procedere
    if stato_attivita:
        stato_attivita = stato_attivita.strip()
    else:
        stato_attivita = "Non disponibile"

    # Formattazione dei dati estratti con il Codice ATECO
    company_data = f"Nome Azienda: {company_name}\nIndirizzo: {company_address}\nCodice CCRea: {company_code_rea}\nPartita IVA: {partita_iva}\nStato Attività: {stato_attivita}\nCod_Ateco: {ateco_value}\n"

    progress_var.set(progress_var.get() + 1)

    if stato_attivita == "Attiva":
        active_text.insert(tk.END, f"{company_name}\n")
        active_count_var.set(active_count_var.get() + 1)
        return company_data
    else:
        inactive_text.insert(tk.END, f"{company_name} (Non attiva)\n")
        inactive_count_var.set(inactive_count_var.get() + 1)
        return None

def get_next_page_link(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Solleva un'eccezione per errori HTTP
    except requests.exceptions.RequestException as e:
        print(f"Errore durante la richiesta HTTP: {str(e)}")
        return None

    selector = parsel.Selector(text=response.text)
    next_page_link = selector.css('ul.pagination li.page-item.active + li.page-item a.page-link::attr(href)').get()
    return next_page_link

def scrape_company_info(url, ateco_value, progress_var, active_text, inactive_text, active_count_var, inactive_count_var):
    next_page_url = url
    companies_data = []
    total_pages = 1  # Stima iniziale

    while next_page_url:
        try:
            response = requests.get(next_page_url)
            response.raise_for_status()  # Solleva un'eccezione per errori HTTP
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Errore", f"Errore durante la richiesta HTTP: {str(e)}")
            return []

        selector = parsel.Selector(text=response.text)
        company_links = selector.css('.content-azienda-nome a::attr(href)').getall()

        for company_link in company_links:
            company_url = urljoin(next_page_url, company_link)
            company_data = scrape_company_data(company_url, ateco_value, progress_var, active_text, inactive_text, active_count_var, inactive_count_var)
            if company_data:
                fatturato = extract_fatturato_from_page(requests.get(company_url).text)
                if fatturato:
                    company_data += f"Fatturato: {fatturato}\n"

                # Aggiungi i dati della società alla lista companies_data
                companies_data.append(company_data)

        next_page_link = get_next_page_link(next_page_url)
        next_page_url = urljoin(url, next_page_link) if next_page_link else None

        # Estima il numero di pagine
        if len(companies_data) > 0:
            total_pages = (len(companies_data) // len(company_links)) + 1

    messagebox.showinfo("Estrazione completata", "Estrazione completata.")
    return companies_data

def save_to_excel(data_list, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Aggiungi intestazioni
    headers = ["Nome Azienda", "Indirizzo", "Codice CCRea", "Partita IVA", "Stato Attività", "Cod_Ateco", "Fatturato"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        # Imposta lo stile del testo a nero
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(size=14, bold=True, color="000000")  # Colore nero

    # Aggiungi solo dati delle aziende attive
    row_num = 2  # Inizia dalla seconda riga
    for data in data_list:
        if "Stato Attività: Attiva" in data:  # Verifica se l'azienda è attiva
            data_split = data.split("\n")
            for col_num, header in enumerate(headers):
                # Cerca la sottostringa corrispondente all'header desiderato
                for value in data_split:
                    if header in value:
                        # Estrai il valore dopo i due punti
                        cell_value = value.split(":")[1].strip()
                        # Inserisci il valore nella cella
                        ws.cell(row=row_num, column=col_num + 1, value=cell_value)
                        break
            row_num += 1

    # Salva il file Excel
    wb.save(f"{filename}.xlsx")
    messagebox.showinfo("Salvataggio completato", f"Il file {filename}.xlsx è stato creato con successo!")

def save_to_csv(data_list, filename):
    headers = ["Nome Azienda", "Indirizzo", "Codice CCRea", "Partita IVA", "Stato Attività", "Cod_Ateco", "Fatturato"]
    
    with open(f"{filename}.csv", 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        csvwriter.writerow(headers)
        for data in data_list:
            if "Stato Attività: Attiva" in data:  # Aggiungi solo dati delle aziende attive
                row = []
                data_split = data.split("\n")
                for header in headers:
                    for value in data_split:
                        if header in value:
                            cell_value = value.split(":")[1].strip()
                            row.append(cell_value)
                            break
                csvwriter.writerow(row)

    messagebox.showinfo("Salvataggio completato", f"File CSV salvato come {filename}.csv")

def load_csv_to_database(csv_file):
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="yourpassword",
            database="yourdatabase"
        )
        cursor = conn.cursor()

        with open(csv_file, 'r', encoding='utf-8') as file:
            csv_data = csv.reader(file)
            headers = next(csv_data)

            for row in csv_data:
                cursor.execute("INSERT INTO yourtable (column1, column2, column3, ...) VALUES (%s, %s, %s, ...)", row)

            conn.commit()
            messagebox.showinfo("Caricamento completato", "Caricamento nel database completato con successo!")

    except mysql.connector.Error as e:
        messagebox.showerror("Errore durante il caricamento nel database", f"Errore durante il caricamento nel database: {str(e)}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def main():
    root = tk.Tk()
    root.title("AtecoExtractor")

    # Frame per l'input
    input_frame = tk.Frame(root, padx=10, pady=10)
    input_frame.grid(row=0, column=0, sticky="nsew")

    # Frame per l'output
    output_frame = tk.Frame(root, padx=10, pady=10)
    output_frame.grid(row=1, column=0, sticky="nsew")

    # Frame per le aziende attive
    active_frame = tk.Frame(output_frame, padx=5, pady=5, bg="black")
    active_frame.grid(row=0, column=0, sticky="nsew")

    # Frame per le aziende non attive
    inactive_frame = tk.Frame(output_frame, padx=5, pady=5, bg="black")
    inactive_frame.grid(row=0, column=1, sticky="nsew")

    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Etichetta per spiegare il funzionamento del programma
    explanation_label = tk.Label(input_frame, text="Questo programma consente di estrarre informazioni sulle aziende corrispondenti a un determinato codice ATECO.")
    explanation_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

    # Etichette e campi di input per codice ATECO e nome del file Excel
    ateco_label = tk.Label(input_frame, text="Inserisci il codice ATECO (formato: XX_X):")
    ateco_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
    ateco_entry = tk.Entry(input_frame)
    ateco_entry.grid(row=1, column=1, padx=5, pady=5)

    filename_label = tk.Label(input_frame, text="Nome del file Excel:")
    filename_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
    filename_entry = tk.Entry(input_frame)
    filename_entry.grid(row=2, column=1, padx=5, pady=5)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(input_frame, orient="horizontal", length=300, mode="determinate", variable=progress_var)
    progress_bar.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky="we")

    # Etichette per i conteggi di aziende attive e non attive
    active_count_var = tk.IntVar()
    active_count_label = tk.Label(active_frame, textvariable=active_count_var, font=("Helvetica", 12, "bold"), fg="green", bg="black")
    active_count_label.grid(row=0, column=0, sticky="nsew")

    inactive_count_var = tk.IntVar()
    inactive_count_label = tk.Label(inactive_frame, textvariable=inactive_count_var, font=("Helvetica", 12, "bold"), fg="green", bg="black")
    inactive_count_label.grid(row=0, column=0, sticky="nsew")

    # Titoli per le caselle di testo delle aziende attive e non attive
    active_title_label = tk.Label(active_frame, text="Aziende Attive", font=("Helvetica", 14, "bold"), fg="green", bg="black")
    active_title_label.grid(row=1, column=0, sticky="nsew")

    inactive_title_label = tk.Label(inactive_frame, text="Aziende Non Attive", font=("Helvetica", 14, "bold"), fg="green", bg="black")
    inactive_title_label.grid(row=1, column=0, sticky="nsew")

    # Casella di testo per le aziende attive
    active_text = tk.Text(active_frame, height=20, width=60, wrap="word", bg="black", fg="white")
    active_text.grid(row=2, column=0, sticky="nsew")

    # Casella di testo per le aziende non attive
    inactive_text = tk.Text(inactive_frame, height=20, width=60, wrap="word", bg="black", fg="white")
    inactive_text.grid(row=2, column=0, sticky="nsew")

    # Funzione per estrarre i dati e salvare il file Excel
    def extract_data():
        ateco_code = ateco_entry.get()
        filename = filename_entry.get()

        if not filename:
            messagebox.showwarning("Nome file mancante", "Inserisci un nome per il file Excel.")
            return

        url_ateco = f"https://www.companyreports.it/ateco/{ateco_code}"

        # Pulisce la casella di testo prima di ogni nuova estrazione
        active_text.delete("1.0", tk.END)
        inactive_text.delete("1.0", tk.END)
        progress_var.set(0)
        active_count_var.set(0)
        inactive_count_var.set(0)

        # Esegui l'estrazione dei dati su un thread separato per non bloccare l'interfaccia
        def run_extraction():
            try:
                data_list = scrape_company_info(url_ateco, ateco_code, progress_var, active_text, inactive_text, active_count_var, inactive_count_var)
                if data_list:
                    save_to_excel(data_list, filename)
                    # Caricamento dei dati nel database se l'utente lo desidera
                    if messagebox.askyesno("Caricare nel database?", "Vuoi caricare i dati nel database?"):
                        save_to_csv(data_list, filename)
                        load_csv_to_database(f"{filename}.csv")
                else:
                    messagebox.showinfo("Nessuna azienda trovata", "Nessuna azienda trovata")
            except Exception as e:
                messagebox.showerror("Errore durante l'estrazione", f"Errore durante l'estrazione: {str(e)}")

        # Esegui l'estrazione su un thread separato
        thread = threading.Thread(target=run_extraction)
        thread.start()

    # Pulsante per avviare l'estrazione dei dati
    extract_button = tk.Button(input_frame, text="Estrai dati", command=extract_data)
    extract_button.grid(row=4, column=0, columnspan=2, padx=5, pady=5)

    root.mainloop()

if __name__ == "__main__":
    main()
